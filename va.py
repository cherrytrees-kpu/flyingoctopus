#Standard library modules
import json
import sys
import pathlib
import shutil
import requests
import time
import datetime
#Third-party modules
from openpyxl import Workbook
from Bio.Seq import Seq
from Bio.Alphabet import IUPAC
#Importing functions
def importHGVS(filename):
    """
    importHGVS - imports HGVS into the program
    Parameters: none
    Return: list containing the imported HGVS ids
    """

    inputfile_open = False
    listHGVS = []

    #Open the file
    while inputfile_open == False:
        try:
            inputfile = open(filename, 'r')
            inputfile_open = True
        except IOError:
            print('File not found.\n')

            filename = input('Re-enter filename: ')

    #Import data into listHGVS
    for line in inputfile:
        listHGVS.append(line.strip('\n'))

    inputfile.close()

    return listHGVS
def importanno(filename):
    """
    importanno - from a file containing json.dumps annotation data, import it as
    a list of dictionaries
    Parameters: filename - name of file to open
    Return: a list of dictionaries of annotation data
    """

    inputfile_open = False
    listdata = []

    #Open the file
    while inputfile_open == False:
        try:
            inputfile = open(filename, 'r')
            inputfile_open = True
        except IOError:
            print('File not found.\n')
            filename = input('Re-enter filename: ')

    #Import the data
    for line in inputfile:
        listdata.append(json.loads(line))

    inputfile.close()

    return listdata
def vcftoHGVS(filename):
    """
    vcftoHGVS - calls the myvariant.info get_hgvs_from_vcf function
    Parameters: None
    Returns: none; outputs a file containing all of the HGVS IDs
    """
    listHGVS = []

    inputfile_open = False

    while inputfile_open == False:
        try:
            listHGVS = list(myvariant.get_hgvs_from_vcf(filename))
            inputfile_open = True
        except IOError:
            print('File not found.\n')
            filename = input('Re-enter filename: ')

    #Accept name input from user
    name = input('Please enter an identifier for your file: ')
    outputfile = open('HGVS_' + name + '.txt', 'w')

    #Write to file
    print('Writing to file...')
    #for idHGVS in listHGVS:
    length = len(listHGVS)
    i = 0
    while i < (length - 1):
        outputfile.write(listHGVS[i])
        outputfile.write ('\n')
        i = i + 1
    outputfile.write(listHGVS[i])
    print('Done.')

    outputfile.close()
#Exporting functions
def outputHGVS(listHGVS, name = ""):
    """
    outputHGVS - outputs a file with the name 'HGVS_*name.txt' containing the
    list of HGVS ids
    Parameters:
    listHGVS - list - contains list ids
    Returns: nothing; outputs to file 'HGVS_*name.txt'

    """
    i = 0

    #Accept name for output file from user
    if name != "":
        name = input('Please enter an identifier for your file: ')
    outputfile = open('HGVS_' + name + '.txt', 'w')

    #Write to file
    while(len(listHGVS)) > index:
        outputfile.write(listHGVS[i])
        if i != (len(listHGVS)-1):
            outputfile.write('\n')
        i = i + 1

    outputfile.close()
def exportanno(listanno, filename):
    """
    exportanno - exports raw annotation data
    Parameters:
    listanno - list - list of dictionaries containing raw JSON data
    filename - string - name of the file to output to
    Returns: none; outputs to file with filename
    """
    #Create output file
    outputfile = open(filename, 'w')

    #For each annotation in the annotation list, dump the JSON data to file
    for anno in listanno:
        outputfile.write(json.dumps(anno))
        #Check to see if the last element has been reached before entering a
        #newline character
        if listanno.index(anno) != (len(listanno)-1):
            outputfile.write('\n')
    outputfile.close()
    print('Export completed.' + '\n')
def writeanno(listanno, name = "annotated_mutations.txt"):
    """
    writeanno - export annotations to file, tab delimited.
    Parameters:
    listanno - list - contains all of the information to be exported
    Return: none; outputs to a file named annotated_mutations.txt
    """
    print('Writing to file...')

    outputfile = open(name, 'w')

    #Output:
    #Write header
    outputfile.write('HGVS' + '\t'
                     + 'RSID' + '\t'
                     + 'VarType' + '\t'
                     + 'gnomADG' + '\t'
                     + 'gnomADE' + '\t'
                     + 'ClinVar' + '\t'
                     + 'Most Severe Consequence' + '\t'
                     + 'Genes Affected' + '\t'
                     + 'Relevant Transcripts' + '\n')

    #Write data
    i = 0
    #Write until but not including last element
    while i < (len(listanno) - 1):

        genelist = []
        relevanttranscripts = []

        #For genelist
        if listanno[i]['genelist'][0] is not None:
            for gene in listanno[i]['genelist']:
                 genelist.append(gene['gene_symbol'])
        #For relevanttranscripts
        if listanno[i]['relevanttranscripts'] is not None:
            for transcript in listanno[i]['relevanttranscripts']:
                relevanttranscripts.append(transcript['transcript_id'])

        outputfile.write(str(listanno[i]['_id']) + '\t'
                         + str(listanno[i]['rsid']) + '\t'
                         + str(listanno[i]['vartype']) + '\t'
                         + str(listanno[i]['gnomADG']) + '\t'
                         + str(listanno[i]['gnomADE']) + '\t'
                         + str(listanno[i]['ClinVar']) + '\t'
                         + str(listanno[i]['MScon']) + '\t'
                         + str(genelist) + '\t'
                         + str(relevanttranscripts) + '\t'
                         )

        if 'brain_expression' in listanno[i]:
            if listanno[i]['brain_expression'] is not None:
                outputfile.write(str(listanno[i]['brain_expression']))

        outputfile.write('\n')

        #Progress Indicator
        if i%1000 == 0:
            print (str(i) + ' out of ' + str(len(listanno)) + ' written...')

        i = i + 1
    #Write last element
    last_genelist = []
    last_relevanttranscripts = []

    if listanno[i]['genelist'][0] is not None:
        for gene in listanno[i]['genelist']:
             last_genelist.append(gene['gene_symbol'])
    if listanno[i]['relevanttranscripts'] is not None:
        for transcript in listanno[i]['relevanttranscripts']:
            last_relevanttranscripts.append(transcript['transcript_id'])

    outputfile.write(str(listanno[i]['_id']) + '\t'
                     + str(listanno[i]['rsid']) + '\t'
                     + str(listanno[i]['vartype']) + '\t'
                     + str(listanno[i]['gnomADG']) + '\t'
                     + str(listanno[i]['gnomADE']) + '\t'
                     + str(listanno[i]['ClinVar']) + '\t'
                     + str(listanno[i]['MScon']) + '\t'
                     + str(last_genelist) + '\t'
                     + str(last_relevanttranscripts) + '\t'
                     )

    if 'brain_expression' in listanno[i]:
        if listanno[i]['brain_expression'] is not None:
            outputfile.write(str(listanno[i]['brain_expression']))

    print('Annotations written to file')

    outputfile.close()
def outputsequence(sequence, id, type, path):
    sequencefile = open(path.as_posix()
                    + '/'
                    + id
                    + '_'
                    + type
                    + '.txt',
                    'w',)
    sequencefile.write(str(sequence))
    sequencefile.close()
#Parsing functions
def dumpvartype(idHGVS):
    """
    dumpvartype - determine the type of variant based on HGVS id
    Parameters: idHGVS - the HGVS id of variant being considered
    Return: vartype
    """
    vartype = ''
    if '>' in idHGVS:
        vartype = 'snv'
    elif 'del' in idHGVS:
        vartype = 'del'
    elif 'ins' in idHGVS:
        vartype = 'ins'
    return vartype
def converthgvs(idHGVS):
    """
    converthgvs - convert a HGVS ID into a format that can be used in a filename
    or directory
    Parameters: idHGVS - the HGVS ID as a string
    Return: a string with the converted HGVS ID
    """
    def findchr(idHGVS):
        if idHGVS[3:5].isnumeric() is True:
            return idHGVS[3:5]
        else:
            return idHGVS[3]
    def findpos(idHGVS, marker):
        pos_start = idHGVS.index(marker) + 1
        pos_end = pos_start + 1
        pos_complete = False

        while pos_complete is False:
            if idHGVS[pos_start:pos_end].isnumeric() is False:
                pos_chr = idHGVS[pos_start:(pos_end-1)]
                pos_complete = True
            else:
                pos_end = pos_end + 1
        return pos_chr
    #For substitution
    if '>' in idHGVS:
        #Identify chromosome number
        num_chr = findchr(idHGVS)
        #Identify position
        pos_chr = findpos(idHGVS, '.')
        #Identify allele 1 and 2
        allele1 = idHGVS[idHGVS.index('>') - 1]
        allele2 = idHGVS[idHGVS.index('>') + 1]
        return (num_chr + '-' + pos_chr + '-' + allele1 + '-' + allele2)

    #For deletion and insertion
    elif 'delins' in idHGVS:
        #Identify chromosome numbers
        num_chr = findchr(idHGVS)
        #Identify position
        pos_chr = findpos(idHGVS, '.')
        #Identify inserted sequence
        seq_delins = idHGVS[idHGVS.index('s') + 1:]
        #Debug
        return(num_chr + '-' + pos_chr + '-delins-' + seq_delins)

    #For deletion
    elif 'del' in idHGVS:
        #Identify chromosome number
        num_chr = findchr(idHGVS)
        #Identify position 1
        if '_' in idHGVS:
            pos1_chr = idHGVS[(idHGVS.index('.') + 1):idHGVS.index('_')]
            #Identify position 2
            pos2_chr = findpos(idHGVS, '_')
            return (num_chr + '-' + str(pos1_chr) + '_' + str(pos2_chr) + '-del')
        else:
            pos_chr = findpos(idHGVS, '.')
            return (num_chr + '-' + str(pos_chr) + '-del')

    #For insertion
    elif 'ins' in idHGVS:
        #Identify chromosome number
        num_chr = findchr(idHGVS)
        #Identify positions
        pos1_chr = idHGVS[(idHGVS.index('.') + 1):idHGVS.index('_')]
        pos2_chr = findpos(idHGVS, '_')
        #Identify inserted sequences
        seq_ins = idHGVS[idHGVS.index('s') + 1:]
        return (num_chr + '-' + str(pos1_chr) + '_' + str(pos2_chr) + '-ins' + seq_ins)
#Filtering functions
def filteraffected(listaffected, listcontrol):
    """
    filteraffected - filters from lists of HGVS IDs of affected and control
                     individuals
    Parameters:
    listaffected - list - contains the list of HGVS ids of affected individuals
    listcontrol - list - contains the list of HGVS ids of unaffected individuals
    Returns: list containing candidate mutations only
    """
    i = 0
    listfiltered = []

    print('Filtering starting...')

    while i < len(listaffected[0]):
        candidate = True

        #Check if HGVS is in all listAffected. If not, set candidate to false
        for a in listaffected:
            if (listaffected[0][i] in a) == False:
                candidate = False

        #Check if HGVS is in listControl. If yes, set candidate to False
        for c in listcontrol:
            if listaffected[0][i] in c:
                candidate = False

        #Write to file
        if candidate == True:
            listfiltered.append(listaffected[0][i].strip('\n'))

        if i%1000 == 0:
            print(str(i) + '/' + str(len(listaffected[0])) + ' completed...')

        i = i + 1

    print('Filtering completed.')
    return listfiltered
def filternodata(listanno):
    list_nodata = []
    list_candidate = []

    for anno in listanno:
        nodata = checknodata(anno)
        if nodata is True:
            list_nodata.append(anno)
        elif nodata is False:
            list_candidate.append(anno)

    return list_nodata, list_candidate
def filterfreq(listanno):
    list_overfreq = []
    list_candidate = []

    for anno in listanno:
        overfreq = checkfreq(anno)
        if overfreq is True:
            list_overfreq.append(anno)
        elif overfreq is False:
            list_candidate.append(anno)

    return list_overfreq, list_candidate
def filtercons(listanno):
    list_irrelevant = []
    list_candidate = []

    for anno in listanno:
        irrelevant = checkcons(anno)
        if irrelevant is True:
            list_irrelevant.append(anno)
        elif irrelevant is False:
            list_candidate.append(anno)

    return list_irrelevant, list_candidate
def filterexpression(listanno):
    list_hpa = []
    list_notbrainexpressed = []
    list_candidate = []

    #Get annotation from Human Protein Atlas
    for anno in listanno:
        if anno['genes'] is not None:
            for gene in anno['genes']:
                anno_hpa = hpa.annotate(gene['gene_id'])
                list_hpa.append(anno_hpa)
                if anno_hpa is not None:
                    gene['RNAbrd'] = anno_hpa['RNA brain regional distribution']
                else:
                    gene['RNAbrd'] = None

    #Filtering
    for anno in listanno:
        notbrainexpressflag = True
        if anno['genes'] is not None:
            for gene in anno['genes']:
                if gene['RNAbrd'] != 'Not detected':
                    notbrainexpressflag = False

        if notbrainexpressflag is True:
            list_notbrainexpressed.append(anno)
        else:
            list_candidate.append(anno)

    #Export and return
    exportanno(list_hpa, 'hpa_annotations.txt')
    return list_notbrainexpressed, list_candidate
def checknodata(anno):
    """
    checknodata - checks if variant has any data
    Parameters: data - variant and its associated annotations
    Returns: True or False
    """
    nodataflag = False
    if anno['MScon'] == 'N/A':
        nodataflag = True

    return nodataflag
def checkfreq(anno):
    """
    checkfreq - checks if variant above the frequency cut off
    Parameters: data - variant and its associated annotations
    Returns: True or False
    """

    overfreqpcflag = False
    if (anno['gnomADG'] != 'N/A') and (anno['gnomADG'] is not None):
        if anno['gnomADG'] >= 0.01:
            overfreqpcflag = True
    if (anno['gnomADE'] != 'N/A') and (anno['gnomADE'] != None):
        if anno['gnomADE'] >= 0.01:
            overfreqpcflag = True

    return overfreqpcflag
def checkcons(anno):
    """
    checkcons - checks if the variant is relevant
    Parameters: data - variant and its associated annotations
    Returns: True or False
    """
    nonrelevantflag = False

    if (anno['MScon'] == 'intron_variant'
        or anno['MScon'] == 'non_coding_transcript_exon_variant'
        or anno['MScon'] == 'upstream_gene_variant'
        or anno['MScon'] == 'downstream_gene_variant'
        or anno['MScon'] == 'synonymous_variant'
        or anno['MScon'] == '5_prime_UTR_variant'
        or anno['MScon'] == '3_prime_UTR_variant'
        or anno['MScon'] == 'intergenic_variant'):
        nonrelevantflag = True

    return nonrelevantflag
#Processing functions
def combineanno(listmvi, listvep, listHGVS):
    """
    combineanno - combine annotations from myvariant.info and VEP
    Parameters:
    listmvi - list containing dictionaries of myvariant.info annotations
    listvep - list containing dictionaries of VEP annotations
    listHGVS - list containing HGVS IDs
    Return: return combined annotation list
    """
    #listanno - holds list of combined annotations
    listanno = []
    i = 0

    #Do for every variant in listHGVS
    while i < len(listHGVS):
        #True if there's data
        anno_mvi_flag = False
        anno_vep_flag = False

        #Check if there's data
        if listmvi[i] is not None:
            anno_mvi_flag = True
        if listvep[i] is not None:
            anno_vep_flag = True

        #Progress tracker
        if i%100 == 0:
            print(str(i) + ' out of ' + str(len(listHGVS)) + ' completed..')

        #Combine annotations
        listanno.append(dict({'_id':listHGVS[i],
                        'rsid':mvi.dumpRSID(listmvi[i]),
                        'vartype':dumpvartype(listHGVS[i]),
                        'gnomADG':mvi.dumpgnomADG(listmvi[i]),
                        'gnomADE':mvi.dumpgnomADE(listmvi[i]),
                        'ClinVar':mvi.dumpCV(listmvi[i]),
                        'MScon':vep.dumpconsequence(listvep[i]),
                        'genes':vep.dumpensembltranscripts(listvep[i])
                        }))
        i = i + 1

    return listanno
def getsequences(listanno, basepath):
    def substitution(transcript, pos, allele, strand):
        if strand == -1:
            allele = str(Seq(allele, IUPAC.unambiguous_dna).complement())
        return transcript[:pos-1] + allele + transcript[pos:]

    def insertion(transcript, start, end, insert, strand):
        if strand == -1:
            insert = str(Seq(insert, IUPAC.unambiguous_dna).complement())
        return transcript[:start] + insert + transcript[end:]

    def deletion(transcript, start, end):
        return transcript[:start-1] + transcript[end:]

    def translation (sequence):
        if 'N' not in sequence:
            dna = Seq(sequence, IUPAC.unambiguous_dna)
            mrna = dna.transcribe()
            protein = mrna.translate()
            return str(protein)
        else:
            print('Ambiguous nucleotide present, no protein')

    #Do for every variant
    for anno in listanno:
        #Holds stats
        list_stats = []
        #Make a folder for variant
        print ('Working on ' + anno['_id'] + '...')
        variantpath = basepath.joinpath(converthgvs(anno['_id']))
        print(variantpath)
        variantpath.mkdir(exist_ok=True)
        #Do for every gene
        for gene in anno['genes']:
            #Stats to track
            list_transcripts = []
            list_notranscripts = []
            #Make folder for gene
            genepath = variantpath.joinpath(gene['gene_id'])
            genepath.mkdir(exist_ok=True)
            #Do for every transcript
            for transcript in gene['transcripts']:
                list_transcripts.append(transcript['transcript_id'])
                print('Pulling transcripts for ' + transcript['transcript_id'])
                cdna = vep.ensemblsequence(transcript['transcript_id'], 'cdna')
                cds = vep.ensemblsequence(transcript['transcript_id'], 'cds')
                #Write cdna and cds
                outputsequence(cdna, transcript['transcript_id'], 'cdna', genepath)
                outputsequence(cds, transcript['transcript_id'], 'cds', genepath)
                #Get rid of splice_region_variant, TF_binding_site, regulatory_region_variant
                if cds is not None:
                    get_protein = True
                else:
                    get_protein = False
                    list_notranscripts.append(transcript['transcript_id'])
                for term in transcript['consequence_terms']:
                    if (term == 'splice_region_variant'
                        or term == 'TF_binding_site'
                        or term == 'regulatory_region_variant'
                        or term == 'splice_donor_variant'):
                        get_protein = False
                #Get protein sequences
                if (get_protein is True) and (anno['vartype'] == 'snv'):
                    vcds = substitution(cds,
                                        transcript['cds_start'],
                                        transcript['variant_allele'],
                                        transcript['strand'],
                                        )
                    vprotein = translation(vcds)
                elif (get_protein is True) and (anno['vartype'] == 'del'):
                    vcds = deletion(cds,
                                    transcript['cds_start'],
                                    transcript['cds_end'],
                                    )
                    vprotein = translation(vcds)
                elif (get_protein is True) and (anno['vartype'] == 'ins' or anno['vartype' == 'delins']):
                    vcds = insertion(cds,
                                    transcript['cds_start'],
                                    transcript['cds_end'],
                                    transcript['variant_allele'],
                                    transcript['strand']
                                    )
                    vprotein = translation(vcds)
                #Write proteins
                if get_protein is True:
                    protein = translation(cds)
                    outputsequence(protein, transcript['transcript_id'], 'protein', genepath)
                    outputsequence(vprotein, transcript['transcript_id'], 'vprotein', genepath)
            #Add to the stats document
            list_stats.append(dict({'gene':gene['gene_id'],
                                    'transcripts':list_transcripts,
                                    'no_transcripts':list_notranscripts,
                                    }))
        #Write the summary file for the variant
        summary_file = open(variantpath.as_posix()
                            + '/'
                            + converthgvs(anno['_id'])
                            + '_summary.txt'
                            ,
                            'w',
                            )
        for item in list_stats:
            summary_file.write(item['gene']
                                + '\n'
                                + 'All transcripts:'
                                + str(item['transcripts'])
                                + '\n'
                                + 'No sequence:'
                                + str(item['no_transcripts'])
                                + '\n'
                                )

        summary_file.close()
def exportexcel(listanno):
    #Variables
    workbook = Workbook()
    worksheet = workbook.active
    #Do for every annotation
    for anno in listanno:
        idHGVS = converthgvs(anno['_id'])
        sheet = workbook.create_sheet(idHGVS[0:31])

        #Variant annotations
        #Write the header
        sheet.cell(row=1, column=1).value = 'HGVS'
        sheet.cell(row=1, column=2).value = 'RSID'
        sheet.cell(row=1, column=3).value = 'Variant Type'
        sheet.cell(row=1, column=4).value = 'gnomADG'
        sheet.cell(row=1, column=5).value = 'gnomADE'
        sheet.cell(row=1, column=6).value = 'ClinVar Reports'
        sheet.cell(row=1, column=7).value = 'Most Severe Consequence'
        #Write Data
        sheet.cell(row=2, column=1).value = anno['_id']
        sheet.cell(row=2, column=2).value = anno['rsid']
        sheet.cell(row=2, column=3).value = anno['vartype']
        sheet.cell(row=2, column=4).value = anno['gnomADG']
        sheet.cell(row=2, column=5).value = anno['gnomADE']
        sheet.cell(row=2, column=6).value = str(anno['ClinVar'])
        sheet.cell(row=2, column=7).value = anno['MScon']

        r_index = 4
        c_index = 1

        #Do for every gene in the annotation
        for gene in anno['genes']:
            #Get Uniprot data for the gene
            data_uniprot = uniprot.annotate(gene['gene_id'])
            c_index = 1
            #Write the gene symbol and gene id
            sheet.cell(row=r_index, column=c_index).value = gene['gene_symbol']
            sheet.cell(row=r_index, column=c_index + 1).value = gene['gene_id']
            r_index = r_index + 1
            if data_uniprot is not None:
                sheet.cell(row=r_index, column=c_index).value = 'Protein name'
                sheet.cell(row=r_index, column=c_index + 1).value = data_uniprot['protein_name']
                sheet.cell(row=r_index, column=c_index + 2).value = data_uniprot['acc']
                r_index = r_index + 1
                sheet.cell(row=r_index, column=c_index).value = 'Function'
                sheet.cell(row=r_index, column=c_index + 1).value = data_uniprot['function']
                r_index = r_index + 1
            #Write the transcript information
            for transcript in gene['transcripts']:
                #Reset column index
                c_index = 1
                #Get the key values
                keys_all = list(transcript.keys())
                keys = []
                #Get desired keys
                keys.append('transcript_id')
                keys.append('strand')
                keys.append('consequence_terms')
                keys.append('variant_allele')
                if 'cdna_start' in keys_all:
                    keys.append('cdna_start')
                    keys.append('cdna_end')
                if 'cds_start' in keys_all:
                    keys.append('cds_start')
                    keys.append('cds_end')
                    keys.append('codons')
                    keys.append('amino_acids')
                    keys.append('protein_start')
                    keys.append('protein_end')
                #Output transcript information
                for key in keys:
                    sheet.cell(row=r_index, column = c_index).value = key
                    sheet.cell(row=r_index + 1, column = c_index).value = str(transcript[key])
                    c_index = c_index + 1
                r_index = r_index + 2
            #Write rest of uniprot information
            if data_uniprot is not None:
                for feature in data_uniprot['features']:
                    keys = list(feature.keys())
                    c_index = 1
                    for key in keys:
                        sheet.cell(row=r_index, column = c_index).value = key
                        sheet.cell(row=r_index + 1, column = c_index).value = str(feature[key])
                        c_index = c_index + 1
                    r_index = r_index + 2

        workbook.save(filename = 'test.xlsx')
